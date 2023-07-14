import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

# Read the Excel file
df = pd.read_excel("data.xlsx", sheet_name="Sheet")

# Convert the date and time columns to datetime objects
df["Date"] = pd.to_datetime(df["Date"])
df["Time"] = pd.to_datetime(df["Time"], format="%H:%M:%S")

# Perform datetime operations
df["Year"] = df["Date"].dt.year
df["Month"] = df["Date"].dt.month
df["Day"] = df["Date"].dt.day
df["Hour"] = df["Time"].dt.hour
df["Minute"] = df["Time"].dt.minute
df["Second"] = df["Time"].dt.second

wb = Workbook()
ws = wb.active

# Append the DataFrame rows to the worksheet
for row in dataframe_to_rows(df, index=False, header=True):
    ws.append(row)

# Set the cell format for the Date and Time columns
date_column = ws["A"]
time_column = ws["B"]

# Save the workbook with the updated data
wb.save("data.xlsx")