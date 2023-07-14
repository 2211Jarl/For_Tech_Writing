import datetime
from openpyxl import Workbook
import time
  
# Creating a workbook object
workbook_obj = Workbook()
  
# Selecting the active sheet
active_sheet = workbook_obj.active
  
# Adding heading to cell 1
active_sheet.cell(row=1, column=1).value = "Data"
active_sheet.cell(row=2, column=1).value = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

time.sleep(4)
  
active_sheet.cell(row=3, column=1).value = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
  
time.sleep(2)

active_sheet.cell(row=4, column=1).value = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
  
# Saving the workbook object 
workbook_obj.save('Tutorials Point.xlsx')
workbook_obj.close()